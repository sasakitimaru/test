from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

ppp_l = ["入間郡三芳町","新座市","狭山市","加須"]#,"入間市","所沢市"]
driver_path = "driver/chromedriver"
driver = webdriver.Chrome(executable_path=driver_path)
driver.implicitly_wait(10)
driver.get("https://www.google.co.jp/search?q=%E5%9F%BC%E7%8E%89%E7%9C%8C%E3%80%80%E9%A3%AF%E8%83%BD%E5%B8%82%E3%80%80%E5%85%AC%E5%9C%92&biw=885&bih=789&tbm=lcl&sxsrf=ALeKk02hSjSqoXN-1AuSKa9kERQdC2sJ1A%3A1620457607874&ei=hziWYNjmNMXM-QafqqzABA&oq=%E5%9F%BC%E7%8E%89%E7%9C%8C%E3%80%80%E9%A3%AF%E8%83%BD%E5%B8%82%E3%80%80%E5%85%AC%E5%9C%92&gs_l=psy-ab.3..35i39k1j38l4.12461.16328.0.17064.16.14.0.0.0.0.262.1666.0j7j2.9.0....0...1c.1j4.64.psy-ab..10.6.1019...0j0i7i4i30k1j0i7i30k1j0i433i131k1j0i13i30k1j0i30k1j35i304i39k1.0.MqCoMLvE_sw#rlfi=hd:;si:;mv:[[35.870376799999995,139.3506587],[35.8283287,139.29228880000002]];tbs:lrf:!1m4!1u3!2m2!3m1!1e1!1m4!1u2!2m2!2m1!1e1!2m1!1e2!2m1!1e3!3sIAE,lf:1,lf_ui:1")


for ppp in ppp_l:
    pname = []
    address_list = []
    time.sleep(2)
    driver.find_element(By.NAME, "q").clear()
    driver.find_element(By.NAME, "q").click()
    driver.find_element(By.NAME, "q").send_keys("埼玉県　"+ppp+"　公園" + Keys.ENTER)
    yy=0
    while True:
        #time.sleep(2)
        c=driver.find_elements_by_class_name("cXedhc")
        x = driver.find_elements(By.CSS_SELECTOR,".rllt__details")
        for i in c:
            try:
                pk = i.find_element(By.CSS_SELECTOR, "span > div:nth-child(1)")
                if "公園" in pk.text:
                    pppname = i.find_element(By.CSS_SELECTOR, "div:nth-child(2) > div")
                    pname.append(pppname.text)
            except:
                pass
        for j in x:
            try:
                pk = j.find_element(By.CSS_SELECTOR, "div:nth-child(1)")
                if "公園" in pk.text:
                    try:
                        y = j.find_element(By.CSS_SELECTOR, "div:nth-child(2)")
                        address_list.append(y.get_attribute("textContent"))
                    except:
                        print("NULL")
                        address_list.append(None)
            except:
                pass
        try:
            #pnameって名前のIDを持つ奴が次へのボタンだった
            driver.find_element(By.ID,"pnnext").click()
            #多分2でも頑張れば1でも上手くいきそう
            time.sleep(3)
        except:
            break
        yy=yy+1
        #最後の方は住所ないやつとかあってずれちゃうから５周（20*5=100箇所）で止めてる
        if yy==8:
            break
    """
    while i < len(pname):
        if address_list[i] is None:
            pass
        else:
            if "埼玉県" in address_list[i]:
                del pname[i]
                del address_list[i]
                i=i-1
        i = i+1
        """
    pname2 = []
    address_list2 = []
    for i in range(len(pname)):
        if address_list[i] is None:
            pname2.append(pname[i])
            address_list2.append(address_list[i])
        else:
            if "埼玉県" in address_list[i]:
                pass
            else:
                pname2.append(pname[i])
                address_list2.append(address_list[i])
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.worksheets[0]
    row=1
    for qw in pname2:
        ws_new.cell(row,1).value = qw
        row=row+1
    row=1
    for we in address_list2:
        ws_new.cell(row,2).value = we
        row=row+1

    for col in ws_new.columns:
        max_length = 0
        column = col[0].column

        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.3
        ws_new.column_dimensions[col[0].column_letter].width = adjusted_width
    wb_new.save(ppp+"＿お試し.xlsx")

#print(c[7].text)
driver.close()

