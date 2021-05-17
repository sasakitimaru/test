from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import xml.etree.ElementTree as ET
import openpyxl
import time
wb_new = openpyxl.load_workbook("八潮市.xlsx")
ws_new = wb_new.worksheets[0]

driver_path = "driver/chromedriver"
driver = webdriver.Chrome(executable_path=driver_path)
driver.implicitly_wait(10)
address_list = []
Parkl = []
for hh in ws_new.iter_rows():
    Parkl.append(hh[1].value)
c=0
for park in Parkl:
    driver.get("https://www.google.co.jp/")
    driver.find_element(By.NAME, "q").click()
    kensaku = "埼玉県　"+ park + "　住所"
    driver.find_element(By.NAME, "q").send_keys(kensaku + Keys.ENTER)
    #s = a.find_element(By.CSS_SELECTOR, "span.GRkHZd w8qArf")
    if park != "朝霞中央公園":
        try:
           # time.sleep(1)
            a = driver.find_element(By.CSS_SELECTOR, "div.sXLaOe")
            address_list.append(a.text)
        except Exception:
            try:
                driver.find_element(By.CSS_SELECTOR, ".uMdZh:nth-child(1) .dbg0pd > span").click()
                #tmp=driver.find_element(By.CSS_SELECTOR,"div.zloOqf PZPZlf")
                sex=driver.find_element(By.CSS_SELECTOR,"span.LrzXr")
                address_list.append(sex.get_attribute("textContent"))
            except:
                sex=driver.find_element(By.CSS_SELECTOR,"span.LrzXr")
                address_list.append(sex.text)
    c=c+1
    print(c)
driver.quit()



row_num=1
for unti in address_list:
    ws_new.cell(row_num,3).value = unti
    row_num = row_num + 1
    
for col in ws_new.columns:
    max_length = 0
    column = col[0].column

    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))

    adjusted_width = (max_length + 2) * 1.3
    ws_new.column_dimensions[col[0].column_letter].width = adjusted_width
wb_new.save("八潮市公園.xlsx")
