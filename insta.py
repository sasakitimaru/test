from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import re
import openpyxl

wb1 = openpyxl.load_workbook("/Users/ookitomoya/documents/test/insta.xlsx")
ws1 = wb1.worksheets[0]
list = []
for roo in ws1.iter_rows():
    if roo[2].value is None:
        break
    list.append(roo[2].value)
driver_path = "/Users/ookitomoya/documents/test/driver/chromedriver"
driver = webdriver.Chrome(executable_path=driver_path)
driver.implicitly_wait(5)
"""
cnt = 0
for roo in ws1.iter_rows():
    for i in range(5,len(roo)+1):
        driver.get(list[cnt])
        a = driver.find_elements(By.CSS_SELECTOR, "span.g47SY")
        if not roo[i] is None:
            roo[i].append(a[1].get_attribute("textContent"))
            break
    cnt = cnt + 1
"""
a_list = []
for i in list:
    #print(i)
    driver.get(i)
    a = driver.find_elements(By.CSS_SELECTOR, "span.g47SY")
    a_list.append(a[1].get_attribute("textContent"))    #print(a[1].get_attribute("textContent"))]
cnt = 0
for roo in ws1.iter_rows():
    if roo[2].value is None:
        break
    for i in range(1,len(roo)):
        s = len(roo)-i
        if roo[s] is None:
            pass
        else:
            break
    print(s)
    ws1.cell(cnt+1,s+2).value = a_list[cnt]
    cnt = cnt + 1
wb1.save("insta.xlsx")
driver.quit()

